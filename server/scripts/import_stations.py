import sys
from datetime import datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

SERVER_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(SERVER_DIR))

from database.connection import Base, SessionLocal, create_database_if_not_exists, engine
from models.station import StationModel


TARGET_STATION_IDS = {
    "00105",
    "00106",
    "00203",
    "00204",
    "00207",
    "04244",
}


def normalize_station_id(value) -> str:
    if value is None:
        raise ValueError("대여소 번호가 비어 있습니다.")
    return str(int(value)).zfill(5)


def clean_text(value, field_name: str) -> str:
    text = str(value).strip() if value is not None else ""
    if not text:
        raise ValueError(f"{field_name} 값이 비어 있습니다.")
    return text


def clean_coordinate(value, field_name: str) -> Decimal:
    if value is None:
        raise ValueError(f"{field_name} 값이 비어 있습니다.")
    return Decimal(str(value))


def clean_date(value):
    if isinstance(value, datetime):
        return value.date()
    if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
        return value
    raise ValueError("설치일자 형식이 올바르지 않습니다.")


def parse_target_rows(file_path: Path) -> dict[str, dict]:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    if "대여소현황" not in workbook.sheetnames:
        raise ValueError("'대여소현황' 시트를 찾을 수 없습니다.")

    worksheet = workbook["대여소현황"]
    stations = {}
    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=6, values_only=True), start=6
    ):
        if row[0] is None:
            continue
        station_id = normalize_station_id(row[0])
        if station_id not in TARGET_STATION_IDS:
            continue

        lcd_rack_count = int(row[7] or 0)
        qr_rack_count = int(row[8] or 0)
        latitude = clean_coordinate(row[4], "위도")
        longitude = clean_coordinate(row[5], "경도")
        if not (Decimal("37") <= latitude <= Decimal("38")):
            raise ValueError(f"{row_number}행 위도가 서울 범위를 벗어났습니다.")
        if not (Decimal("126") <= longitude <= Decimal("128")):
            raise ValueError(f"{row_number}행 경도가 서울 범위를 벗어났습니다.")

        stations[station_id] = {
            "station_id": station_id,
            "station_name": clean_text(row[1], "대여소명"),
            "district": clean_text(row[2], "자치구"),
            "address": clean_text(row[3], "상세주소"),
            "latitude": latitude,
            "longitude": longitude,
            "installed_at": clean_date(row[6]),
            "lcd_rack_count": lcd_rack_count,
            "qr_rack_count": qr_rack_count,
            "rack_count": lcd_rack_count + qr_rack_count,
            "operation_type": clean_text(row[9], "운영 방식"),
        }

    workbook.close()
    missing_ids = TARGET_STATION_IDS - stations.keys()
    if missing_ids:
        raise ValueError(f"Excel에서 대여소를 찾지 못했습니다: {sorted(missing_ids)}")
    return stations


def import_stations(file_path: Path) -> tuple[int, int]:
    station_rows = parse_target_rows(file_path)
    create_database_if_not_exists()
    Base.metadata.create_all(bind=engine)

    inserted = 0
    updated = 0
    with SessionLocal() as db:
        try:
            for station_id, values in station_rows.items():
                station = db.get(StationModel, station_id)
                if station is None:
                    db.add(StationModel(**values))
                    inserted += 1
                else:
                    for key, value in values.items():
                        setattr(station, key, value)
                    updated += 1
            db.commit()
        except Exception:
            db.rollback()
            raise
    return inserted, updated


def main() -> None:
    default_file = SERVER_DIR / "data" / "seoul_bike_data.xlsx"
    file_path = Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else default_file
    if not file_path.exists():
        raise FileNotFoundError(f"Excel 파일을 찾을 수 없습니다: {file_path}")

    inserted, updated = import_stations(file_path)
    print(f"대여소 적재 완료 - 신규: {inserted}개, 수정: {updated}개")


if __name__ == "__main__":
    main()
